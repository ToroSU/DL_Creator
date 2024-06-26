import openpyxl
from openpyxl.styles import Alignment


def set_border(ws, cell_range):
    # border around
    # ref.: https://jingwen-z.github.io/how-to-munipulate-excel-workbook-by-python/
    rows = ws[cell_range]
    side = openpyxl.styles.Side(border_style="thick", color="000000") 

    rows = list(rows)
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = openpyxl.styles.Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def create_list(list_info, os_info, data_input, data_all_input):
    ## Creat list and save as Excel
    list_ver = "v" + list_info[2]
    fn_temp = [list_info[1], os_info[0], os_info[1], "x64", "Drv", list_ver]
    fn_temp = "_".join(fn_temp)
    fn = fn_temp + ".xlsx"

    wb = openpyxl.Workbook()
    wb_1 = wb.create_sheet("The lastest release", 0) # Add worksheet and specify location
    wb_2 = wb.create_sheet("Release note", 0)

    # Define cell color and fill it
    cell_A1_color_fill = openpyxl.styles.fills.PatternFill(patternType="solid", start_color="FFFF00")
    title_color_fill = openpyxl.styles.fills.PatternFill(patternType="solid", start_color="4F81BD")
    blank_color_fill = openpyxl.styles.fills.PatternFill(patternType="solid", start_color="DCE6F1")
    gray_color_fill = openpyxl.styles.fills.PatternFill(patternType="solid", start_color="C0C0C0")
    
    wb_1["A1"].fill = cell_A1_color_fill

    for row in wb_1["A2:W2"]: # This line for fill cell color by column
        for cell in row:
            cell.fill = title_color_fill

    for i in range(0, len(data_input), 2): # Odd column
        for row in wb_1["A{}:W{}".format(str(i+3), str(i+3))]: # +3 for cell column number in Driver list
            for cell in row:
                cell.fill = blank_color_fill

    # Border
    thin = openpyxl.styles.Side(border_style="thin", color="000000") 
    thick = openpyxl.styles.Side(border_style="thick", color="000000") 
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
    thick_border = openpyxl.styles.Border(left=thick, right=thick, top=thick, bottom=thick)

    for row in wb_1["A2:W{}".format(str(len(data_input)+2))]:
        for cell in row:
            cell.border = border

    # Font define
    title_font = openpyxl.styles.Font(size=12, bold=False, name='Arial Unicode MS', color="FFFFFF")
    content_font = openpyxl.styles.Font(size=10, bold=False, name='Verdana', color="000000")
    sheetRN_content_font = openpyxl.styles.Font(size=12, bold=False, name='Verdana', color="000000")

    # Alignment define
    content_align = Alignment(horizontal = "center")
    content_column_non_align_list = [1, 2, 16, 17, 22, 23] # column : [A, B, P, Q, V] do not participate in alignment.
    content_column_non_align_list = [temp - 1 for temp in content_column_non_align_list] # Do -1 for excel format in python for loop.

    # List title at excel cell:"A1"
    cell_A1_title = [str(list_info[1]), "Driver List -", str(os_info[0]), str(os_info[1]), "Version:", str(list_info[2]), "- Update Date:", str(list_info[3])]
    cell_A1_title = " ".join(cell_A1_title)
    wb_1["A1"].value = str(cell_A1_title)
    wb_1["A1"].font = content_font

    # row2_column_title : Each column title at "The lastest release" (row 2)
    row2_column_title = ["Category", "Description", "Provider/ODM", "Vendor", "List of Support model name", "Driver type", "HSA", "APP ID", "APP Name",
                         "Support side link", "Extension ID", "Support OS", "WHQL", "Check version mechanism", "Hardware ID", "Driver version", "Driver Date",
                         "VersionRegKey", "Silent install command", "Match FW Version", "Need reboot(Y/N)", "Driver package", "Remark"]

    for i in range(len(row2_column_title)):
        wb_1.cell(column=i+1, row=2).value = str(row2_column_title[i])
        wb_1.cell(column=i+1, row=2).font = title_font

    # All list to excel, modify font and alignment
    for i in range(0, len(data_all_input)): # i = number of columns.
        wb_1.cell(column=i+1, row=2).alignment = content_align
        for j in range(0, len(data_all_input[i])): # j = number of rows.
            if i in content_column_non_align_list: # column which do not participate in alignment.
                wb_1.cell(column=i+1, row=j+3).value = str(data_all_input[i][j])
                wb_1.cell(column=i+1, row=j+3).font = content_font
            else:
                wb_1.cell(column=i+1, row=j+3).value = str(data_all_input[i][j])
                wb_1.cell(column=i+1, row=j+3).font = content_font
                wb_1.cell(column=i+1, row=j+3).alignment = content_align


    # Set column width
    dims = {}
    for row in wb_1.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        if value <= 20:
            wb_1.column_dimensions[col].width = 20
        else:
            wb_1.column_dimensions[col].width = value + 7

    wb_1.column_dimensions["A"].width = 30 # A column do not automatically adjust column width (Cause A1)


    # create release note sheet (sheetRN) (wb_2)
    os_edition = str(os_info[0]) # e.g. Win11
    os_version = str(os_info[1]) # 22H2
    os_build = str(os_info[2]) # 22621.3
    update_date = str(list_info[3]) # 2022/11/18
    list_version = str(list_info[2]) # 0.01
    subject_str = "First release for {} OS.".format(os_version)

    sheetRN_title_str = "\"Driver List - {}\" Model Release notes".format(os_edition)
    sheetRN_updatedate_str = "Update Date: {}".format(update_date)
    sheetRN_column_title_list = ["Version", "Release Date", "Subject", "OS", "Version", "Remark"]
    sheetRN_column_ini_content_list = [list_version, update_date, subject_str, os_edition, os_build, ""]
    sheetRN_column_width_list = [12, 20, 110, 12, 15, 35]

    for i in range(0, len(sheetRN_column_title_list)):
        wb_2.cell(column=i+1, row=3).value = str(sheetRN_column_title_list[i])
        wb_2.cell(column=i+1, row=3).font = sheetRN_content_font

        wb_2.cell(column=i+1, row=4).value = str(sheetRN_column_ini_content_list[i])
        wb_2.cell(column=i+1, row=4).font = sheetRN_content_font

        wb_2.cell(column=i+1, row=3).alignment = Alignment(horizontal = "center", vertical = "center")
        wb_2.cell(column=i+1, row=4).alignment = Alignment(horizontal = "center", vertical = "center") 

        col = wb_2.cell(column=i+1, row=3).column_letter
        wb_2.column_dimensions[col].width = sheetRN_column_width_list[i]
    
    wb_2.merge_cells("A1:E1") # merge cells
    wb_2["A1"].value = str(sheetRN_title_str)
    wb_2["F1"].value = str(sheetRN_updatedate_str)
    wb_2["A1"].font = sheetRN_content_font
    wb_2["F1"].font = sheetRN_content_font
    wb_2["A1"].alignment = Alignment(horizontal = "center", vertical = "center") 
    wb_2["F1"].alignment = Alignment(horizontal = "right", vertical = "center") 


    for row in wb_2["A3:F3"]: # This line for fill cell color by column
        for cell in row:
            cell.fill = gray_color_fill

    # set border
    for row in wb_2["A3:F4"]:
        for cell in row:
            cell.border = border

    BORDER_LIST = ["A3:F3"]
    for pos in BORDER_LIST: 
        set_border(wb_2, pos)


    # Save to exccel file
    try:
        del wb["Sheet"] # Delete default sheet which name is "Sheet"
        wb.save(fn)
    except:
        print("\nExcel WARNING : Please try closing the excel file and RE-RUN the program.")

