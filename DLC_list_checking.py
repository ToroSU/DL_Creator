import os
import openpyxl
from openpyxl.styles import Alignment
from PyQt5.QtWidgets import QApplication, QMessageBox
import sys
import logging

def excel_reader():
    # 若有複數個 excel檔案，回傳第一個檔名
    excels = []
    dir_path = os.getcwd() 
    files = os.listdir(dir_path)
    try:
        for file in files:
            if ".xlsx" in file:
                excels.append(file)
        return excels[0]

    # 無 Driver list file, 強制結束
    except:
        app = QApplication(sys.argv)
        QMessageBox.about(None, "WARNING", "No Driver List in current path")
        return 0
    
    
def version_checking(ver_1, ver_2):
    ver_1_split = ver_1.split(".")
    ver_2_split = ver_2.split(".")
    try:
        ver_checking_status = 0
        for i in range(0, len(ver_1_split)):
            if int(ver_1_split[i]) != int(ver_2_split[i]):
                ver_checking_status += 1

    except:
        print("error in version checking")

    if ver_checking_status == 0:
        ckecking_status_return = True
    else:
        ckecking_status_return = False

    return ckecking_status_return


def set_all_row_font(font_input, max_col, index_i, sheet_driver_list):
    for i in range(0, max_col):
        sheet_driver_list.cell(row=index_i+1, column=i+1).font = font_input


def excel_typesetting_and_save(current_sheet, workbook, excel_file_name):
    # Set column width
    dims = {}
    for row in current_sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        if value <= 20:
            current_sheet.column_dimensions[col].width = 20
        else:
            current_sheet.column_dimensions[col].width = value + 7

    current_sheet.column_dimensions["A"].width = 30 # A column do not automatically adjust column width (Cause A1)

    # Set alignment
    # Alignment define
    content_align_vertical = Alignment(wrap_text = True, vertical = "center")
    content_align_center = Alignment(horizontal = "center", vertical= "center")
    content_column_non_align_horizontal_list = [1, 2, 15, 16, 17, 22, 23] # column : [A, B, P, Q, V] do not participate in alignment horizontal.
    content_column_non_align_horizontal_list = [temp - 1 for temp in content_column_non_align_horizontal_list] # Do -1 for excel format in python for loop.

        # All list to excel, modify font and alignment
    for i in range(0, current_sheet.max_column): # i = number of columns.
        current_sheet.cell(column=i+1, row=2).alignment = content_align_center # row2 (title row)
        for j in range(0, current_sheet.max_row): # j = number of rows.
            if i in content_column_non_align_horizontal_list: # column which do not participate in alignment.
                current_sheet.cell(column=i+1, row=j+3).alignment = content_align_vertical
            else:
                current_sheet.cell(column=i+1, row=j+3).alignment = content_align_center

    #save
    workbook.save(excel_file_name)


#另外用一個file reader 來開 sys_inf_check_list.txt 編碼問題 機歪
def file_reader(file_):
    # Open the .inf file. If open faild, try to use "utf-16" decode and ignore the error.
    try:  #utf-8 & utf-16 
        f = open(file_, encoding="utf-8")
        lines_ = f.readlines()
        f.close()
    except:
        try :
            f = open(file_, encoding="iso_8859_1")  # ANSI
            # print(file_, " not utf-8 encode")
            lines_ = f.readlines()
            f.close()
        except :
            f = open(file_, encoding="utf-16-le") # UTF16-LE 
            lines_ = f.readlines()
            f.close()

    return lines_


def get_date_ver(date_ver_line):
    content_ = date_ver_line.split(":")[1]
    content_ = content_.strip(" ")
    content_ = content_.strip("\n")
    date_return, ver_return = content_.split(" ")

    return date_return, ver_return


def get_description(description_line):
    content_ = description_line.split(":")[1]
    content_ = content_.strip("\n")
    description_return = content_.strip(" ")

    return description_return


def get_hardwareID(hardwareID_line):
    if "Hardware IDs:" in hardwareID_line or "硬體識別碼" in hardwareID_line:
        content_ = hardwareID_line.split(":")[1]
        content_ = content_.strip("\n")
        hardwareID_return = content_.strip(" ")
    else:
        content_ = hardwareID_line.strip("\n")
        hardwareID_return = content_.strip(" ")
    
    return hardwareID_return


def list_checking_main():
    ## Main start
    sys_inf_chk_file = "syschecklist.txt"

    # create sys inf check list
    # TODO from system and from package can use it
    os.system('cmd /c "pnputil.exe /enum-devices /drivers /ids >{}"'.format(sys_inf_chk_file))
    print("Creating...")

    if os.path.isfile(sys_inf_chk_file):
        print("Create inf check list successfully")

    else:
        print("No file: {} in current path".format(sys_inf_chk_file))
        return 0

    date_column = 17 # 定位各column位置，萬一未來要改會好改一點點
    version_column = 16
    hardwardID_column = 15
    description_column = 2

    lines = file_reader(sys_inf_chk_file)

    print(lines[0])
    if "公用程式" in lines[0]:
        print("The current client system is Chinese")
        instanceID_str = "執行個體識別碼"
        originalName_str = "原始名稱"
        driverVersion_str = "驅動程式版本"
        deviceDesc_str = "裝置描述"
        hardwareIDs_str = "硬體識別碼"
        compatibleIDs = "相容識別碼"
        matchingDrivers = "相符的驅動程式"

    else:
        instanceID_str = "Instance ID"
        originalName_str = "Original Name"
        driverVersion_str = "Driver Version"
        deviceDesc_str = "Device Description"
        hardwareIDs_str = "Hardware IDs"
        compatibleIDs = "Compatible IDs"
        matchingDrivers = "Matching Drivers"
    
    ins_ID_index = []

    # 紀錄 Instance_ID 的位置
    for i in range(0, len(lines)):
        if instanceID_str in lines[i]:
            ins_ID_index.append(i)

    # 將 sys_inf_check_list 分成一個個block
    block_list = []
    for i in range(0, len(ins_ID_index)-1):
        temp_list = lines[ins_ID_index[i]:ins_ID_index[i+1]-1] # e.g. lines[2:24-1]
        block_list.append(temp_list)

    # 篩選出含有 OEM inf 及 origin inf 的block
    # 將不包含關鍵字的block存成index，再倒序刪除
    index_for_del = [x for x in range(0, len(block_list))] # 預備索引[0 ~ block 數量] 的list

    for i in range(0, len(block_list)): 
        for line in block_list[i]:
            if originalName_str in line: # 有Original Name 代表他有 oem ，至少目前觀察到是這樣
                index_for_del.remove(i)
                break

    
    # 刪除無original name block
    index_for_del.reverse() # 索引倒序
    for i in index_for_del:
        del block_list[i]


    # 開讀 excel
    excel_file_name = excel_reader()
    print(excel_file_name)
    rd_wb = openpyxl.load_workbook(excel_file_name)

    sheet_release_note = rd_wb[rd_wb.sheetnames[0]]
    sheet_driver_list = rd_wb[rd_wb.sheetnames[1]] # [1] list sheet 目前用這個就好

    max_row = sheet_driver_list.max_row  
    max_col = sheet_driver_list.max_column

    # 讀取 remark col, 暫存 inf的地方
    remark_column_rows = []
    for row_i in range(0, max_row):
        temp = sheet_driver_list.cell(row = row_i+1, column = 23)
        remark_column_rows.append(temp.value)


    # 找 driver list 中存在的inf 及其對應的指標 match_inf_index
    match_inf_index = [] # index of inf match block (excel cell)
    match_block_index = [] # index of inf match block (block)
    have_one_match = False
    # 由於match_inf_index 是基於 driver list excel，所以迴圈從2開始, 若第一列就有對應inf，就會從2開始 e.g. [2, 3, 5, 7]，
    # 得到 "有在系統中找到的inf" ， 與block的對應關係 (match_inf, match_block) 
    for i in range(2, len(remark_column_rows)-1):
        current_inf = remark_column_rows[i]
        for block_i in range(0, len(block_list)):
            
            for block_i_line in block_list[block_i]:
                if current_inf.lower() in block_i_line.lower():
                    have_one_match = True
                    break

                else:
                    have_one_match = False

            if have_one_match:
                match_inf_index.append(i)
                match_block_index.append(block_i)
                have_one_match = False
                break


    ## 在有找到inf的前提下(match_inf_index)，尋找description/HWID/Date/Ver
    date_list = []
    ver_list = []
    description_list = []
    hardwardID_list = []

    for i in range(0, len(match_inf_index)): # 先遍歷 remark_column_rows ，索引為有對應到系統inf的項目 match_inf，也就是把excel中有對到系統的Inf一個個抓出來
        index_i = match_inf_index[i] # match_inf_index:[2, 4, 5, 6] => index_i: 2,...
        current_inf = remark_column_rows[index_i]
        # print(current_inf)

        find_current_inf = False # 讀到有 current_inf 後才開始偵測 "Driver Version" ，才會是剛好對應的 version/date
        for block_line_i in range(0, len(block_list[match_block_index[i]])):
            block_index_i = match_block_index[i]

            if current_inf.lower() in block_list[block_index_i][block_line_i].lower():  
                # sub_block_start = block_line_i
                find_current_inf = True # 讀到有 current_inf 後才開始偵測 "Driver Version" ，才會是剛好對應的 version/date

            if find_current_inf and driverVersion_str in block_list[block_index_i][block_line_i]:
                date_version_line = block_list[block_index_i][block_line_i]
                find_current_inf = False
            
            if deviceDesc_str in block_list[block_index_i][block_line_i]:
                description_line = block_list[block_index_i][block_line_i]

            # 這邊取第二條 HWID，注意某些地方可能只有一條，如此她們的下一行應該會有"Matching Drivers:" or "Compatible IDs:"，寫個判斷式避開就好
            if hardwareIDs_str in block_list[block_index_i][block_line_i]:
                if compatibleIDs in block_list[block_index_i][block_line_i + 1] or matchingDrivers in block_list[match_block_index[i]][block_line_i+1]:
                    hardwareID_line = block_list[block_index_i][block_line_i]
                else:
                    hardwareID_line = block_list[block_index_i][block_line_i + 1]

        # 以下list皆與match_inf_index(index_i) 對應，為在找到inf前提下所存下的索引
        date_str, ver_str = get_date_ver(date_version_line)
        description_str = get_description(description_line)
        hardwardID_str = get_hardwareID(hardwareID_line)

        date_list.append(date_str)
        ver_list.append(ver_str)
        description_list.append(description_str)
        hardwardID_list.append(hardwardID_str)


    # font define
    checking_fail_font = openpyxl.styles.Font(size=10, bold=False, name='Verdana', color="FF0000") # red
    checking_success_font = openpyxl.styles.Font(size=10, bold=False, name='Verdana', color="00B050") # green

    ## 在有找到inf的前提下(match_inf_index)，date/Ver 錯誤 標紅色
    # 找到的description / HWID 填入指定格子內
    # 若 excel cell 中已有相同 dexcription/ HWID 不填入, 若無，添加進去
    for i in range(0, len(match_inf_index)):
        index_i = match_inf_index[i] # 在 list 中的指標
        current_inf = remark_column_rows[index_i]

        # 來自系統的Ground truth
        checking_date = date_list[i] 
        checking_ver = ver_list[i]
        checking_description = description_list[i]
        checking_hardwareID = hardwardID_list[i]

        # 來自 driver list 的答案
        date_in_driver_list_cell = sheet_driver_list.cell(row = index_i+1, column = date_column).value
        ver_in_driver_list_cell = sheet_driver_list.cell(row = index_i+1, column = version_column).value
        description_in_driver_list_cell = sheet_driver_list.cell(row = index_i+1, column = description_column).value
        hardwareID_in_driver_list_cell = sheet_driver_list.cell(row = index_i+1, column = hardwardID_column).value

        # checking date and version
        ver_check_status = version_checking(ver_in_driver_list_cell, checking_ver) # avoid check bug: 03.00.00.1457 and 3.0.0.1457
        if date_in_driver_list_cell == checking_date and ver_check_status: 
            set_all_row_font(checking_success_font, max_col, index_i, sheet_driver_list) # 設定第 index_i 列顏色
        else:
            set_all_row_font(checking_fail_font, max_col, index_i, sheet_driver_list)
        
        # checking description and fill it
        if description_in_driver_list_cell != None:
            temp_split = description_in_driver_list_cell.split("\n")
            if checking_description not in temp_split:
                # 若無重複，即添加一行
                temp_string = description_in_driver_list_cell + checking_description + "\n"
                sheet_driver_list.cell(row=index_i+1, column=description_column).value = str(temp_string)
                # sheet_driver_list.cell(row=index_i+1, column=description_column).alignment = Alignment(wrapText=True)
                # sheet_driver_list.cell(row=index_i+1, column=description_column).alignment = Alignment(vertical= "center")
                sheet_driver_list.cell(row=index_i+1, column=description_column).alignment = Alignment(wrapText=True)
            else:
                pass
        else:
            temp_string = checking_description + "\n"
            sheet_driver_list.cell(row=index_i+1, column=description_column).value = str(temp_string)


        # checking hardwareID and fill it
        if hardwareID_in_driver_list_cell != None:
            temp_split = hardwareID_in_driver_list_cell.split("\n")
            if checking_hardwareID not in temp_split:
            # 若無重複，即添加一行
                temp_string = hardwareID_in_driver_list_cell + checking_hardwareID + "\n" 
                sheet_driver_list.cell(row=index_i+1, column=hardwardID_column).value = str(temp_string)
                sheet_driver_list.cell(row=index_i+1, column=hardwardID_column).alignment = Alignment(wrapText=True)
                # sheet_driver_list.cell(row=index_i+1, column=hardwardID_column).alignment = Alignment(horizontal = "center", vertical= "center")
        else:
            temp_string = checking_hardwareID + "\n"
            sheet_driver_list.cell(row=index_i+1, column=hardwardID_column).value = str(temp_string)

    rd_wb.save(excel_file_name)

    try:
        os.remove(sys_inf_chk_file)
    except OSError as e:
        print(e)
    else:
        print("The directory is deleted successfully")


# 測試中: 多設一個欄位"Check counter"讓list可以持續被檢查，不會因為換sku，之前測過的list還被改變顏色
def list_checking_main_test():
    FORMAT = '%(asctime)s %(levelname)s: %(message)s' # for logging setting
    logging.basicConfig(level=logging.DEBUG, filename='Checking.log', filemode='a', format=FORMAT)
    ## Main start
    sys_inf_chk_file = "syschecklist.txt"

    # create sys inf check list
    # TODO from system and from package can use it
    os.system('cmd /c "pnputil.exe /enum-devices /drivers /ids >{}"'.format(sys_inf_chk_file))
    print("Creating...")

    if os.path.isfile(sys_inf_chk_file):
        print("Create inf check list successfully")

    else:
        print("No file: {} in current path".format(sys_inf_chk_file))
        return 0

    date_column = 17 # 定位各column位置，萬一未來要改會好改一點點
    version_column = 16
    hardwardID_column = 15
    description_column = 2
    checkCounter_column = 24 # just for tool use

    lines = file_reader(sys_inf_chk_file)

    instanceID_str = "Instance ID"
    originalName_str = "Original Name"
    driverVersion_str = "Driver Version"
    deviceDesc_str = "Device Description"
    hardwareIDs_str = "Hardware IDs"
    compatibleIDs = "Compatible IDs"
    matchingDrivers = "Matching Drivers"
    
    ins_ID_index = []

    # 紀錄 Instance_ID 的位置
    for i in range(0, len(lines)):
        if instanceID_str in lines[i]:
            ins_ID_index.append(i)

    # 將 sys_inf_check_list 分成一個個block
    block_list = []
    for i in range(0, len(ins_ID_index)-1):
        temp_list = lines[ins_ID_index[i]:ins_ID_index[i+1]-1] # e.g. lines[2:24-1]
        block_list.append(temp_list)

    # 篩選出含有 OEM inf 及 origin inf 的block
    # 將不包含關鍵字的block存成index，再倒序刪除
    index_for_del = [x for x in range(0, len(block_list))] # 預備索引[0 ~ block 數量] 的list

    for i in range(0, len(block_list)): 
        for line in block_list[i]:
            if originalName_str in line: # 有Original Name 代表他有 oem ，至少目前觀察到是這樣
                index_for_del.remove(i)
                break

    
    # 刪除無original name block
    index_for_del.reverse() # 索引倒序
    for i in index_for_del:
        del block_list[i]


    # 開讀 excel
    excel_file_name = excel_reader()
    print(excel_file_name)
    rd_wb = openpyxl.load_workbook(excel_file_name)

    sheet_release_note = rd_wb[rd_wb.sheetnames[0]]
    sheet_driver_list = rd_wb[rd_wb.sheetnames[1]] # [1] list sheet 目前用這個就好

    max_row = sheet_driver_list.max_row  
    max_col = sheet_driver_list.max_column

    # 讀取 remark col, 暫存 inf的地方
    remark_column_rows = []
    for row_i in range(0, max_row):
        temp = sheet_driver_list.cell(row = row_i+1, column = 23)
        remark_column_rows.append(temp.value)


    # 找 driver list 中存在的inf 及其對應的指標 match_inf_index
    match_inf_index = [] # index of inf match block (excel cell)
    match_block_index = [] # index of inf match block (block)
    have_one_match = False
    # 由於match_inf_index 是基於 driver list excel，所以迴圈從2開始, 若第一列就有對應inf，就會從2開始 e.g. [2, 3, 5, 7]，
    # 得到 "有在系統中找到的inf" ， 與block的對應關係 (match_inf, match_block) 
    for i in range(2, len(remark_column_rows)-1):
        current_inf = remark_column_rows[i]
        try:
            for block_i in range(0, len(block_list)):
                for block_i_line in block_list[block_i]:
                    if current_inf.lower() in block_i_line.lower():
                        have_one_match = True
                        break

                    else:
                        have_one_match = False

                if have_one_match:
                    match_inf_index.append(i)
                    match_block_index.append(block_i)
                    have_one_match = False
                    break
        except:
            print(current_inf) # bug waive : search until none 
            break


    ## 在有找到inf的前提下(match_inf_index)，尋找description/HWID/Date/Ver
    date_list = []
    ver_list = []
    description_list = []
    hardwardID_list = []


    for i in range(0, len(match_inf_index)): # 先遍歷 remark_column_rows ，索引為有對應到系統inf的項目 match_inf，也就是把excel中有對到系統的Inf一個個抓出來
        index_i = match_inf_index[i] # match_inf_index:[2, 4, 5, 6] => index_i: 2,...
        current_inf = remark_column_rows[index_i]
        # print(current_inf)

        find_current_inf = False # 讀到有 current_inf 後才開始偵測 "Driver Version" ，才會是剛好對應的 version/date
        for block_line_i in range(0, len(block_list[match_block_index[i]])):
            block_index_i = match_block_index[i]

            if current_inf.lower() in block_list[block_index_i][block_line_i].lower():  
                # sub_block_start = block_line_i
                find_current_inf = True # 讀到有 current_inf 後才開始偵測 "Driver Version" ，才會是剛好對應的 version/date

            if find_current_inf and driverVersion_str in block_list[block_index_i][block_line_i]:
                date_version_line = block_list[block_index_i][block_line_i]
                find_current_inf = False
            
            if deviceDesc_str in block_list[block_index_i][block_line_i]:
                description_line = block_list[block_index_i][block_line_i]

            # 這邊取第二條 HWID，注意某些地方可能只有一條，如此她們的下一行應該會有"Matching Drivers:" or "Compatible IDs:"，寫個判斷式避開就好
            if hardwareIDs_str in block_list[block_index_i][block_line_i]:
                if compatibleIDs in block_list[block_index_i][block_line_i + 1] or matchingDrivers in block_list[match_block_index[i]][block_line_i+1]:
                    hardwareID_line = block_list[block_index_i][block_line_i]
                else:
                    hardwareID_line = block_list[block_index_i][block_line_i + 1]

        # 以下list皆與match_inf_index(index_i) 對應，為在找到inf前提下所存下的索引
        date_str, ver_str = get_date_ver(date_version_line)
        description_str = get_description(description_line)
        hardwardID_str = get_hardwareID(hardwareID_line)

        date_list.append(date_str)
        ver_list.append(ver_str)
        description_list.append(description_str)
        hardwardID_list.append(hardwardID_str)


    # font define
    checking_fail_font = openpyxl.styles.Font(size=10, bold=False, name='Verdana', color="FF0000") # red
    checking_success_font = openpyxl.styles.Font(size=10, bold=False, name='Verdana', color="00B050") # green

    ## 在有找到inf的前提下(match_inf_index)，且Check counter 為 none (意即從未被檢查過)若date/Ver 錯誤 標紅色
    # 若date/Ver 正確 標綠色
    # 只要有找到inf 對應的check counter+1
    # 找到的description / HWID 填入指定格子內
    # 若 excel cell 中已有相同 dexcription/ HWID 不填入, 若無，添加進去
    for i in range(0, len(match_inf_index)): 
        index_i = match_inf_index[i] # 在 list 中的指標
        current_inf = remark_column_rows[index_i]

        # 來自系統的Ground truth
        checking_date = date_list[i] 
        checking_ver = ver_list[i]
        checking_description = description_list[i]
        checking_hardwareID = hardwardID_list[i]

        # 來自 driver list 的答案
        category_str = sheet_driver_list.cell(row = index_i+1, column = 1).value # for logging message
        date_in_driver_list_cell = sheet_driver_list.cell(row = index_i+1, column = date_column).value
        ver_in_driver_list_cell = sheet_driver_list.cell(row = index_i+1, column = version_column).value
        description_in_driver_list_cell = sheet_driver_list.cell(row = index_i+1, column = description_column).value
        hardwareID_in_driver_list_cell = sheet_driver_list.cell(row = index_i+1, column = hardwardID_column).value
        checkCounter_in_driver_list_cell = sheet_driver_list.cell(row = index_i+1, column = checkCounter_column).value

        if checkCounter_in_driver_list_cell != None:
            temp = checkCounter_in_driver_list_cell.split(":")[1]
            checkCounter = int(temp)
        else:
            checkCounter = 0

        temp_string = "Check Counter:{}".format(checkCounter +1)
        # checking date and version
        ver_check_status = version_checking(ver_in_driver_list_cell, checking_ver) # avoid check bug: 03.00.00.1457 and 3.0.0.1457
        if date_in_driver_list_cell == checking_date and ver_check_status: 
            set_all_row_font(checking_success_font, max_col, index_i, sheet_driver_list) # 設定第 index_i 列顏色
            sheet_driver_list.cell(row=index_i+1, column=checkCounter_column).value = str(temp_string)

        else:
            dateVerChecking_message = "{0} Date/Ver error, Inf Name:{1}, Expect:{2}{3}, Detected:{4}{5}".format(category_str, current_inf, date_in_driver_list_cell, ver_in_driver_list_cell, checking_date, checking_ver)
            logging.warning(dateVerChecking_message)
            if checkCounter == 0:
                set_all_row_font(checking_fail_font, max_col, index_i, sheet_driver_list) # 若 Date ver 不對，但checkcounter == 0 (Ori None)  
        
        
        
        # checking description and fill it
        if description_in_driver_list_cell != None:
            temp_split = description_in_driver_list_cell.split("\n") # e.g. ["Desc.1"] ~ ["Desc.1", "Desc.2", "Desc.3"]        
            if checking_description not in temp_split:
                # 若無重複，即添加一行
                temp_split.append(checking_description)
                temp_string = "\n".join(temp_split)
                sheet_driver_list.cell(row=index_i+1, column=description_column).value = str(temp_string)
            else:
                pass
        else:
            temp_string = checking_description
            sheet_driver_list.cell(row=index_i+1, column=description_column).value = str(temp_string)


        # checking hardwareID and fill it
        if hardwareID_in_driver_list_cell != None:
            temp_split = hardwareID_in_driver_list_cell.split("\n") # e.g. ["HWID1"] ~ ["HWID1", "HWID2", "HWID3"] 
            if checking_hardwareID not in temp_split:
            # 若無重複，即添加一行
                temp_split.append(checking_hardwareID)
                temp_string = "\n".join(temp_split) 
                sheet_driver_list.cell(row=index_i+1, column=hardwardID_column).value = str(temp_string)
                # sheet_driver_list.cell(row=index_i+1, column=hardwardID_column).alignment = Alignment(wrapText=True)
        else:
            temp_string = checking_hardwareID
            sheet_driver_list.cell(row=index_i+1, column=hardwardID_column).value = str(temp_string)

    # save
    # rd_wb.save(excel_file_name)
    excel_typesetting_and_save(sheet_driver_list, rd_wb, excel_file_name)
    logging.info("Checking Complete.\n")

    try:
        os.remove(sys_inf_chk_file)
    except OSError as e:
        print(e)
    else:
        print("The directory is deleted successfully")